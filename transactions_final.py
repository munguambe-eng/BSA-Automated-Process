# %%

import datetime
import pandas as pd
import pymssql
import logging
import openpyxl as px
import win32com.client
import pathlib


# %%
import pandas 
print(pandas.__file__)

# %%
# lista de holidays em Moçambique para 2024
from datetime import date 
year=date.today().year
holidays = [
    datetime.date(year, 1, 1),  # Dia da Paz e Reconciliação
    datetime.date(year, 2, 3),  # Dia dos Heróis Moçambicanos
    datetime.date(year, 4, 7),  # Dia da Mulher Moçambicana
    datetime.date(year, 5, 1),  # Dia do Trabalhador
    datetime.date(year, 6, 25),  # Dia da Independência Nacional
    datetime.date(year, 9, 7),  # Dia da Vitória
    datetime.date(year, 9, 25),  # Dia das Forças Armadas de Libertação Nacional
    datetime.date(year, 10, 4),  # Dia da Paz e da Reconciliação
    datetime.date(year, 12, 25),  # Dia de Natal
]


# %%
def its_holiday(dia):
    return dia in holidays

# %%
def report_date(todays_date):
    if (todays_date-datetime.timedelta(days=1)).weekday()==6: #segunda feira
        if its_holiday(todays_date-datetime.timedelta(days=3)):
            report_date=todays_date-datetime.timedelta(days=4)
            # return report_date,'='
            return report_date,'='+"'"+str(report_date)+"'"
        elif its_holiday(todays_date-datetime.timedelta(days=4)):
            report_date=todays_date-datetime.timedelta(days=4)
            report_date1=todays_date-datetime.timedelta(days=3)
            return report_date1,' between '+"'"+str(report_date)+"'"+' and '+"'"+str(report_date1)+"'"
            # return todays_date-datetime.timedelta(days=4),'and following day, i must review the query and accomodate this scenario'
        else:
            report_date=todays_date-datetime.timedelta(days=3)
            # return todays_date-datetime.timedelta(days=3),'='
            return report_date,'= '+"'"+str(report_date)+"'"
    elif todays_date.weekday()==1:
        if its_holiday(todays_date-datetime.timedelta(days=1)):
            report_date=todays_date-datetime.timedelta(days=4)
            # return todays_date-datetime.timedelta(days=4),'='
            return report_date,'= '+"'"+str(report_date)+"'"
        else:
            if its_holiday(todays_date-datetime.timedelta(days=4)):
                report_date=todays_date-datetime.timedelta(days=4)
                # return todays_date-datetime.timedelta(days=4),'>='
                return report_date,'>= '+"'"+str(report_date)+"'"
            else:
                report_date=todays_date-datetime.timedelta(days=3)
                reportar=todays_date-datetime.timedelta(days=1)
                # return todays_date-datetime.timedelta(days=3),'>='
                return reportar,'>= '+"'"+str(report_date)+"'"
    elif todays_date.weekday()==2:
        if its_holiday(todays_date-datetime.timedelta(days=2)):
            report_date=todays_date-datetime.timedelta(days=4)
            
            # return todays_date-datetime.timedelta(days=4),'>='
            return report_date,'>='+ "'"+str(report_date)+"'"
        elif its_holiday(todays_date-datetime.timedelta(days=1)):
            report_date=todays_date-datetime.timedelta(days=2)
            # return todays_date-datetime.timedelta(days=2),'='
            return report_date,'= '+"'"+str(report_date)+"'"
        else:    
            report_date=todays_date-datetime.timedelta(days=1)
            # return todays_date-datetime.timedelta(days=1),'='
            return report_date,'= '+"'"+str(report_date)+"'"
    elif todays_date.weekday()==3:
        if its_holiday(todays_date-datetime.timedelta(days=1)):
            report_date=todays_date-datetime.timedelta(days=2)
            # return todays_date-datetime.timedelta(days=2),'='
            return report_date,'= '+"'"+str(report_date)+"'"
        else:
            report_date=todays_date-datetime.timedelta(days=1)
            # return todays_date-datetime.timedelta(days=1),'='
            return report_date,'= '+"'"+str(report_date)+"'"
        
    elif todays_date.weekday()==4:
        if its_holiday(todays_date-datetime.timedelta(days=1)):
            report_date=todays_date-datetime.timedelta(days=2)
            # return todays_date-datetime.timedelta(days=2),'='
            return report_date,'= '+"'"+str(report_date)+"'"
        elif its_holiday(todays_date-datetime.timedelta(days=2)):
            report_date=todays_date-datetime.timedelta(days=2)
            report_date1=todays_date-datetime.timedelta(days=1)
            # return todays_date-datetime.timedelta(days=2),'and following day, i must review the query and accomodate this scenario'
            return report_date1, ' between '+"'"+str(report_date)+"'"+' and '+"'"+str(report_date1)+"'"
        else:
            report_date=todays_date-datetime.timedelta(days=1)
            # return todays_date-datetime.timedelta(days=1),'='            ,
            return report_date,'= '+"'"+str(report_date)+"'"
     
def logConfigurations():
    # Get current date details
    today = datetime.datetime.now()
    logFilename = "Log - "+today.strftime("%Y-%m-%d")+".log"

    logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s",
                        filename=logFilename, level=logging.DEBUG)

    console = logging.StreamHandler()
    console.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    console.setFormatter(formatter)
    logging.getLogger().addHandler(console)


def getFromDatabase():
    
  
        
# Change dates on this section if necessary and on fill def

    today = datetime.date.today()
    # today=datetime.date(2026,2,9)    
    current_date,operator=report_date(today)
    # current_date,operator = today,between
    # current_date=datetime.datetime.strptime(current_date,'%Y-%m-%d')
    
    logging.info("Getting transaction data")
    print(current_date)

    # today=current_date.strftime('%Y-%m-%d')
    # current_date = datetime.datetime.now()
    # yesterday = current_date - datetime.timedelta(days=3)
    # today=yesterday.strftime('%Y-%m-%d')

    today="'"+str(today)+"'"
    query = open("query_final.sql", "r").read()
    print(today,'date')
    query = query.format(operator,operator,operator,operator)
    print(query)

    user = "SBICMZ01\\sa001970"
    # user = "SBICMZ01\\EA241478"
    # password = "ke:jVyfpjq2k"
    password = "ht3D9vZroLlujlt>"

    conn = pymssql.connect(server='PMOZ-PA01DRLSNR.mz.sbicdirectory.com',
                           user=user, password=password, database='RSVR_RAW_LAYER_REL')

    #cursor = conn.cursor()
    # cursor.execute(query)
    #data = cursor.fetchall()

    df = pd.read_sql_query(query, conn)
    print(df.columns)
    credit=df[df['Tipo de CartÃ£o']=='Credito']
    if credit.empty:
        print('No credit cards transactions, Dataframe is Empty')
        # raise SystemExit
    df.columns
    df=df.dropna(subset=['Data da liquidaÃ§Ã£o'])
    df.shape
    df.to_excel('original_report.xlsx')

    df=df.drop('ft',axis=1)
    # df=df.drop('Numero Único do Cliente',axis=1)
    
    
    df.shape
    print(df)

    conn.close()

    return df

def getDFfromCSV(reportpath):
    logging.info("Reading CSV file at: "+reportpath)
    df = pd.read_csv(reportpath)
    df.to_excel('cards.xlsx')
    print(df)
    print(reportpath)
    return df


def fillExcelTemplate(df):
    # Change date here too
    logging.info("Get info from DataFrame to Excel File Template")
    today =  datetime.date.today()
    # today=datetime.date(2026,2,9)
    # yesterday = today - datetime.timedelta(days=1)
    # if yesterday.weekday() == 6:
    # if yesterday was a Sunday, get the date for the last Friday
        # last_friday = yesterday - datetime.timedelta(days=2)
    # print the year, month, and day of the last Friday
        # print(last_friday)
        # print("Last Friday's year:", last_friday.year)
        # print("Last Friday's month:", last_friday.month)
        # print("Last Friday's day:", last_friday.day)
        # today=last_friday
    # else:
        # today=yesterday

    today,operador=report_date(today)
    # today,operator = today,between
    # today=datetime.datetime.strptime(today,'%Y-%m-%d')
    print(today.strftime('%Y-%m-%d'),'vamos ver')
    filename = 'Mapa_Controlo de Transacoes _Exterior_Cartoes.xlsx'
    excelworkbook = px.load_workbook(filename)
    sheet = excelworkbook.active

    firstRow = 12
    firstCol = 3
    currentCol = 3
    currentRow = 12
    sheet['D5'] = today.strftime('%Y')
    sheet['D6'] = today.strftime('%Y-%m-%d')
    sheet['D7'] = today.strftime('%Y-%m-%d')
    for index, row in df.iterrows():
        for i in range(17):
            if(str(row[i]) != 'nan' and str(row[i]) != 'None'):
                sheet.cell(row=currentRow,
                           column=currentCol).value = str(row[i])
            else:
                sheet.cell(row=currentRow, column=currentCol).value = ''
            currentCol = currentCol + 1
        currentRow = currentRow + 1
        currentCol = firstCol
    

    column=sheet['O']
    for cell in column[11:sheet.max_row+1]:
        if cell.value is not None:
            try:
                cell.value=float(cell.value)
            except(ValueError,TypeError):
                continue

    column=sheet['P']
    for cell in column[11:]:
        if cell.value is not None:
            try:
                cell.value=float(cell.value)
            except(ValueError,TypeError):
                continue
    column=sheet['Q']
    for cell in column[11:]:
        if cell.value is not None:
            try:
                cell.value=float(cell.value)
            except(ValueError,TypeError):
                continue
            
    import time
    column=sheet['R']
    for cell in column[11:]:
        if cell.value is not None:
            try:
                cell_value=datetime.datetime.strptime(cell.value,'%Y-%m-%d %H:%M:%S')
                cell.value=cell_value.strftime('%Y-%m-%d')

            except(ValueError,TypeError):
                continue
    column=sheet['S']
    for cell in column[11:]:
        if cell.value is not None:
            try:
                cell_value=datetime.datetime.strptime(cell.value,'%Y-%m-%d %H:%M:%S')
                cell.value=cell_value.strftime('%Y-%m-%d')
            except(ValueError,TypeError):
                continue
    # for row in sheet.iter_rows():
        # if row[11].value is None:
            # sheet.delete_rows(row[0].row)
    # sheet['N'].number_format='Number'
    # sheet['O'].number_format='Number'
    # 
    # for row in range(12, sheet.max_row+1):
        # N="N"+str(row)
        # O="O"+str(row)
        # sheet[N].number_format = 'Number'
        # sheet[O].number_format = 'Number'

    logging.info("Saving new excel file with Cards Transactions Info...")
    currentPath = str(pathlib.Path().resolve())
    newFilename = "Mapa_Controlo de Transacoes _Exterior_Cartoes_" + \
        today.strftime("%Y%m%d")+".xlsx"
    logging.info("New file name: "+newFilename)
    excelworkbook.save(newFilename)
    excelworkbook.close()
    logging.info('File saved.')

    return newFilename


def sendWithAttEmail(to, cc, subject, body, attachments):
    outlook = win32com.client.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.to = to
    mail.cc = cc
    mail.Subject = subject
    mail.HtmlBody = body
    for attachment in attachments:
        mail.attachments.Add(attachment)
    mail.send
from email import message
from email.message import MIMEPart
from string import Template
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from typing import Mapping
from email.mime.application import MIMEApplication
import os
import logging


class Smtp_mail(object):
    def __init__(self,_mail_from, _mail_pw, _mail_host:str = "zarelay.mail.standardbank.com", _mail_port:int = 25):
        self.mail_host = _mail_host
        self.mail_port = _mail_port
        self.mail_from = _mail_from
        self.mail_pw = _mail_pw
        return

    def send_mail(self, mail_to, content, cc=None, dict_map:dict = None, subject="", attach=None):
        service = smtplib.SMTP(self.mail_host, self.mail_port)
        
        try:
            service.starttls()
            service.login(self.mail_from, self.mail_pw)
            logging.info(f"SMTP server: authentication passed")
        except Exception as e:
            logging.warning(f"SMTP server: authentication failed. Cause {e}")
            pass

        # setup the parameters of the message
        mail_object = MIMEMultipart()
        mail_object['From']=self.mail_from
        mail_object['To']=mail_to
        mail_object['Subject']=subject
        if not cc == None:
            mail_object['CC']=cc 

        #composing body structure
        try:
            with open(content, mode='r', encoding='utf-8') as f:
                f_content = f.read()
                message = Template(f_content).safe_substitute(dict_map)
        except Exception as e:
            message = content

        # add in the message body
        mail_object.attach(MIMEText(message, 'html'))
        if not attach == None:
            if type(attach) != list:
                with open(attach, 'rb') as f:
                    file = MIMEApplication(f.read(), name=os.path.basename(attach))
                    file['Content-Disposition'] = f'attachment; \
                    filename="{os.path.basename(attach)}"'
                    mail_object.attach(file)
            else:
                for filename in attach:
                    with open(filename, 'rb') as f:
                        file = MIMEApplication(f.read(), name=os.path.basename(filename))
                        file['Content-Disposition'] = f'attachment; \
                        filename="{os.path.basename(filename)}"'
                        mail_object.attach(file)

        # send the message via the server set up earlier.
        try:
            service.send_message(mail_object)
        except:
            logging.error("SMTP server: Failure in sending "+mail_to)
        logging.info(f"SMTP server: Message sent sucessfully")
        return


def main():
    logConfigurations()
    logging.info("Starting script...")
    df = getFromDatabase()
    #df = getDFfromCSV('cardsReport.csv')
    filename = fillExcelTemplate(df)
    original_final='original_report.xlsx'
    to="CEO-MANUTENCAO-EXCEPCAO@mail.standardbank.com"
    # cc='rodrigues.munguambe@standardbank.co.mz;mwayi.sulila@standardbank.co.mz'
    # to='rodrigues.munguambe@standardbank.co.mz'
    # cc='rodrigues.munguambe@standardbank.co.mz' #mwayi.sulila@standardbank.co.mz' ,tassio.rosario@standardbank.co.mz'
    cc='rodrigues.munguambe@standardbank.co.mz;ornelle.nhaca@standardbank.co.mz;rui.guirrugo@standardbank.co.mz;tassio.rosario@standardbank.co.mz;alter.xavier@standardbank.co.mz'
    # cc='rodrigues.munguambe@standardbank.co.mz;bruno.mazive@standardbank.co.mz;Leonardo.Dias@standardbank.co.mz;Bania.Fonseca@standardbank.co.mz;Douglas.Pindula@standardbank.co.mz;ornelle.nhaca@standardbank.co.mz;alter.xavier@standardbank.co.mz;rui.guirrugo@standardbank.co.mz'
    # cc='Humberto.dimande@standardbank.co.mz;silvia.matos@standardbank.co.mz;osvaldo.cumbucane@standardbank.co.mz;machava.machava@standardbank.co.mz;rui.muchanga@standardbank.co.mz;mwayi.sulila@standardbank.co.mz;Rui.Guirrugo@standardbank.co.mz;marcel.saraiva@standardbank.co.mz'
    subject="Relatório de Transacções no Exterior de Cartões"
    # subject="Automation test, now from scheduler"
    body="Viva Colegas!, \n em anexo o Relatório de transações de cartões no estrangeiro "
    Smtp_mail(_mail_from="rodrigues.munguambe@standardbank.co.mz",_mail_pw="").send_mail(mail_to=to,cc=cc,content=body,subject=subject,attach=[filename,original_final])
    # sendWithAttEmail(to="rui.guirrugo@standardbank.co.mz", cc='Humberto.dimande@standardbank.co.mz;silvia.matos@standardbank.co.mz;osvaldo.cumbucane@standardbank.co.mz;machava.machava@standardbank.co.mz;rui.muchanga@standardbank.co.mz;mwayi.sulila@standardbank.co.mz;Rui.Guirrugo@standardbank.co.mz',
    #                  subject="Relatório de Transacções no Exterior de Cartões",
    #                  body="Anexo o Relatório", attachments=[filename])

# %%
hoje = datetime.datetime.now()
if(hoje not in holidays):
    main()

# %%


# %% [markdown]
# 


